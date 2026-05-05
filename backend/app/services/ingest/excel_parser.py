import openpyxl
import os
from typing import List, Dict, Any

class ExcelParser:
    def __init__(self):
        pass
    
    def parse(self, file_path):
        """
        解析Excel文件，提取文本内容
        """
        text = ""
        blocks: List[Dict[str, Any]] = []
        chunk_index = 0
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            for sheet_name in workbook.sheetnames:
                text += f"Sheet: {sheet_name}\n"
                sheet = workbook[sheet_name]
                
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"
                    blocks.append(
                        {
                            "text": row_text,
                            "page": None,
                            "chunk_index": chunk_index,
                            "sheet": sheet_name,
                            "row_index": row_index,
                        }
                    )
                    chunk_index += 1
                text += "\n"
        except Exception as e:
            print(f"Excel解析错误: {str(e)}")
        
        return {
            "text": text,
            "metadata": {
                "file_name": os.path.basename(file_path),
                "file_size": os.path.getsize(file_path),
                "num_sheets": len(workbook.sheetnames) if 'workbook' in locals() else 0
            },
            "blocks": blocks,
        }
